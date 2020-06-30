#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :get_excel_data.py
# @Time      :2020/6/9 0:00
# @Author    :江梅
from openpyxl import load_workbook
from tools.config_file import ReadConfig
from tools.do_path import GetPath
from tools.get_datas import GetDatas
from tools.do_mysql import GetDataFromMysql
from tools.get_my_log import Get_MyLog

#从exce获取手机号码()
phone_num = GetDatas.Phone
# print(type(phone_num))
#从数据库获取手机号码
# phone_num =int(GetDataFromMysql.get_data_from_mysql('SELECT max(MobilePhone) FROM member')[0])+1
# phone_num = 13811138002

# print(phone_num)

'''读取excel的测试数据'''
class GetDataFromExcel:
    #初始化函数用于传递文件名和表名称
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    #获取excel测试用例的表头
    def get_header_from_excel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_column = sheet.max_column
        excel_header = []
        for j in range(1,max_column+1):
            cell_data = sheet.cell(1,j).value
            excel_header.append(cell_data)
        return excel_header

    #获取测试用例的数据
    def get_data_from_excel(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        header = self.get_header_from_excel()
        excel_test_datas = []
        for i in range(2,max_row+1):
            datas={}
            for j in range(1,max_column+1):
                datas[header[j-1]] = sheet.cell(i,j).value
            excel_test_datas.append(datas)   #获取所有的测试数据
        # print(type(excel_test_datas))
        return excel_test_datas

    def get_test_datas(self):
        # 获取配置文件的数据,利用反射
        res_config = ReadConfig().read_config(GetPath.Config_Path,
                                              'MODE', 'mode_test')
        host_config = ReadConfig().read_config(GetPath.Config_Path, 'HOST', 'host')
        # leve_money_mode = ReadConfig.read_config(GetPath.Config_Path,'CHECK_LEVEL_MONEY','leve_money_mode')
        excel_test_datas = self.get_data_from_excel()
        if eval(res_config)[self.sheet_name] == 'all':
            for item in excel_test_datas:
                if item['data'].find('${tel_1}') != -1:
                    item['data'] = item['data'].replace('${tel_1}', str(phone_num))
                elif item['data'].find('${tel_2}') != -1:
                    item['data'] = item['data'].replace('${tel_2}', str(phone_num - 1))
                elif item['data'].find('${normal_tel}') != -1:
                    item['data'] = item['data'].replace('${normal_tel}', str(getattr(GetDatas,'normal_tel')))
                elif item['data'].find('${add_memberId}') != -1:
                    item['data'] = item['data'].replace('${add_memberId}', str(getattr(GetDatas,'add_memberId')))
                elif item['data'].find('${admin_tel}') != -1:
                    item['data'] = item['data'].replace('${admin_tel}', str(getattr(GetDatas,'admin_tel')))
                elif item['data'].find('${vip_tel}') != -1:
                    item['data'] = item['data'].replace('${vip_tel}', str(getattr(GetDatas,'vip_tel')))
                elif item['data'].find('${memberId}') != -1:
                    item['data'] = item['data'].replace('${memberId}', str(getattr(GetDatas,'memberId')))
                else:
                    item['data'] = item['data']
            for item in excel_test_datas:
                if item['url'].find('${host_1}') != -1:
                    item['url'] = item['url'].replace('${host_1}', host_config)
            return excel_test_datas
        else:
            fina_test_datas = []
            for data in excel_test_datas:
                if data['case'] in eval(res_config)[self.sheet_name]:
                    fina_test_datas.append(data)
            for item in fina_test_datas:
                if item['data'].find('${tel_1}') != -1:
                    item['data'] = item['data'].replace('${tel_1}', str(phone_num))
                elif item['data'].find('${tel_2}') != -1:
                    item['data'] = item['data'].replace('${tel_2}', str(phone_num - 1))
                elif item['data'].find('${normal_tel}') != -1:
                    item['data'] = item['data'].replace('${normal_tel}', str(getattr(GetDatas, 'normal_tel')))
                elif item['data'].find('${add_memberId}') != -1:
                    item['data'] = item['data'].replace('${add_memberId}', str(getattr(GetDatas, 'add_memberId')))
                elif item['data'].find('${admin_tel}') != -1:
                    item['data'] = item['data'].replace('${admin_tel}', str(getattr(GetDatas, 'admin_tel')))
                elif item['data'].find('${vip_tel}') != -1:
                    item['data'] = item['data'].replace('${vip_tel}', str(getattr(GetDatas, 'vip_tel')))
                elif item['data'].find('${memberId}') != -1:
                    item['data'] = item['data'].replace('${memberId}', str(getattr(GetDatas, 'memberId')))
                else:
                    item['data'] = item['data']
            for item in fina_test_datas:
                if item['url'].find('${host_1}') != -1:
                    item['url'] = item['url'].replace('${host_1}', host_config)
            for item in fina_test_datas:
                if item['select_sql'] != None:
                    if item['select_sql'].find('${vip_tel}') != -1:
                        item['select_sql'] = item['select_sql'].replace('${vip_tel}', str(getattr(GetDatas, 'vip_tel')))
                    else:
                        item['select_sql'] = item['select_sql']['sql']
                else:
                    Get_MyLog().info('没有select_sql')

            return fina_test_datas
    def write_back_result(self,row,result_value,Test_Result,amount_result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row+1,sheet.max_column-2).value = result_value
        sheet.cell(row + 1, sheet.max_column-1).value = Test_Result
        sheet.cell(row + 1, sheet.max_column).value = amount_result
        wb.save(self.file_name)

    #更新Excel里面的电话号码：把心获取到的手机号码存到init表的（2,1）位置
    # def set_phone_nume(self,phone_value):
    #     wb = load_workbook(self.file_name)
    #     sheet = wb['init']
    #     sheet.cell(2,1).value = phone_value
    #     wb.save(self.file_name)



if __name__ == "__main__":

   # res1 = GetDataFromExcel(Get_Path.get_path('test_datas','test01.xlsx'),'login').get_data_from_excel()
   res2 = GetDataFromExcel(GetPath.Excel_Path_02, 'recharge').get_test_datas()
   # # print('+++++++++res1+++++++++++++++{}'.format(res1))
   print('+=============res2=============={}'.format(res2))


